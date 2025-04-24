import logging
import os
from pathlib import Path

import openpyxl
from celery import shared_task
from django.core.exceptions import ValidationError
from email_validator import EmailNotValidError
from email_validator import validate_email

from scaleos.geography import models as geography_models
from scaleos.hr import models as hr_models
from scaleos.notifications import models as notification_models
from scaleos.organizations import models as organization_models
from scaleos.shared.functions import birthday_cleaning
from scaleos.shared.functions import mobile_phone_number_cleaning
from scaleos.users import models as user_models

logger = logging.getLogger(__name__)

RESENGO_HEADERS = [
    "First_Name",
    "Last_Name",
    "Email",
    "EmailCount",
    "EmailError",
    "NOEmailErrors",
    "Gender",
    "SiteLanguage",
    "BirthDay",
    "Address",
    "PostalCode",
    "City",
    "Country",
    "Telephone",
    "Mobile",
    "Company",
    "Company_Address",
    "LastActivityDate",
    "Company_PostalCode",
    "Company_City",
    "Company_Country",
    "Company_Email",
    "Company_VAT",
    "LogonTimes",
    "NOVisits",
]


def read_sheet_and_validate_resengo_excel(excel_file_path):
    logger.info("Excpecting Headers: %s", RESENGO_HEADERS)

    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active  # Get the first sheet

    errors = []

    for idx, column in enumerate(sheet.iter_cols()):
        xls_current_header = column[0].value
        logger.info("Current xls header: %s", xls_current_header)
        current_excpected_header = RESENGO_HEADERS[idx]
        logger.info("Current excpected header: %s", current_excpected_header)
        if xls_current_header != current_excpected_header:
            errors.append(
                f"The header structure is not correct. \
                We expect {current_excpected_header} but got {xls_current_header}",
            )
    if errors:
        msg = f"The excel file was not imported correctly \
            because of the following reasons: {errors}"
        raise ValidationError(msg)

    return sheet


def import_resengo_row(row, organization_id=None, force_overwrite=None):  # noqa: PLR0915
    name = row[0].value  # 'Oele',
    family_name = row[1].value  # 'Geirnaert',
    email_address = row[2].value  # 'oelegeirnaert@hotmail.com',
    # row[3].value  # 'EmailCount', # noqa: ERA001
    # row[4].value  # 'EmailError', # noqa: ERA001
    # row[5].value  # 'NOEmailErrors', # noqa: ERA001
    # row[6].value  # 'Gender', # noqa: ERA001
    # person_language = row[7].value  # 'NL', # noqa: ERA001
    birthday = row[8].value  # '2/26/1974',
    street_and_number = row[9].value  # 'Fossebaan 174',
    postal_code = row[10].value  # '1741',
    city = row[11].value  # 'Wambeek',
    country = row[12].value  # 'BE',
    # row[13].value  # 'Telephone', # FIXED # noqa: ERA001
    mobile_phone = row[14].value  # '+32-0485517786, # MOBILE
    # row[15].value  # 'Company', # noqa: ERA001
    # row[16].value  # 'Company_Address', # noqa: ERA001
    # row[17].value  # '12/25/24 12:00', # noqa: ERA001
    # row[18].value  # 'Company_PostalCode', # noqa: ERA001
    # row[19].value  # 'Company_City', # noqa: ERA001
    # row[20].value  # 'Company_Country', # noqa: ERA001
    # row[21].value  # 'Company_Email', # noqa: ERA001
    # row[22].value  # 'Company_VAT', # noqa: ERA001
    # row[23].value  # 'LogonTimes', # noqa: ERA001
    # row[24].value  # 'NOVisits' # noqa: ERA001

    # Basic validation
    if len(email_address) == 0:
        msg = "Without email, we cannot create a user. Row: {row[0].row}"
        logger.warning(msg)
        raise ValidationError(msg)

    try:
        # Check that the email address is valid. Turn on check_deliverability
        # for first-time validations like on account creation pages (but not
        # login pages).
        emailinfo = validate_email(
            email_address,
            check_deliverability=False,
        )

        # After this point, use only the normalized form of the email address,
        # especially before going to a database query.
        email_address = emailinfo.normalized

    except EmailNotValidError as e:
        # The exception message is human-readable explanation of why it's
        # not a valid (or deliverable) email address.
        logger.warning(e)
        raise ValidationError(e) from e

    # Check if a user with that email already exists

    user, user_created = user_models.User.objects.get_or_create(
        email=email_address,
    )
    if not user_created and not force_overwrite:
        msg = "Email address '%s' already exists in row: %s", email_address, row[0].row
        logger.warning(msg)
        raise ValidationError(msg)

    # DATA CLEANING
    # Convert date to correct format if necessary
    mobile_phone = mobile_phone_number_cleaning(mobile_phone)
    birthday = birthday_cleaning(birthday)

    person, person_created = hr_models.Person.objects.get_or_create(
        user_id=user.pk,
    )
    if person_created:
        logger.info("Person created")
    person.name = name
    person.family_name = family_name
    person.birthday = birthday
    logger.info(person.name)
    logger.info(person.family_name)
    person.save()

    person_address, person_address_created = (
        hr_models.PersonAddress.objects.get_or_create(person_id=person.pk)
    )
    if person_address_created:
        logger.info("PersonAddress created")

    if person_address.address is None:
        if street_and_number and postal_code and city and country:
            new_address = geography_models.Address.objects.create()
            new_address.street = street_and_number
            new_address.postal_code = postal_code
            new_address.city = city
            new_address.country = country
            new_address.save()
            person_address.address = new_address
            person_address.save()
        else:
            msg = "We cannot create an address because we do not have all the data"
            logger.warning(msg)
            raise ValidationError(msg)

    organization_customer, organization_customer_created = (
        organization_models.OrganizationCustomer.objects.get_or_create(
            organization_id=organization_id,
            person_id=person.pk,
        )
    )
    if organization_customer_created:
        logger.info("OrganizationCustomer created")

    return True


@shared_task
def import_resengo_excel(
    excel_file_path,
    organization_id=None,
    requested_by_user_id=None,
    force_overwrite=None,
):
    errors = []

    logger.info("Starting excel import task from file: %s...", excel_file_path)
    if organization_id is None:
        logger.error("Without organization id, we cannot import a resengo excel")
        return False

    if not organization_models.Organization.objects.filter(id=organization_id).exists():
        logger.error("The organization with id '%s' does not exist", organization_id)
        return False

    try:
        sheet = read_sheet_and_validate_resengo_excel(excel_file_path)
    except ValidationError as e:
        logger.warning("Error: %s", e)
        return f"Error: {e}"

    row_count = sheet.max_row
    for idx, row in enumerate(sheet.iter_rows(min_row=2)):  # Skip the header row
        logger.info("importing row %s of %s", idx, row_count)

        try:
            import_resengo_row(row, organization_id, force_overwrite)
        except ValidationError as e:
            errors.append(e)

    logger.info("import ready...")

    if os.path.exists(excel_file_path):  # noqa: PTH110
        logger.info("Deleting excel file: %s", excel_file_path)
        Path.unlink(excel_file_path)

    msg = ""
    if not errors:
        msg = "Excel file imported successfully!"
        result = notification_models.Notification.NotificationResult.SUCCESS
        logger.info(msg)
    else:
        msg = "Excel import failed!"
        result = notification_models.Notification.NotificationResult.FAILED
        logger.error(
            "The excel file was not imported correctly \
                    because of the following reasons: %s",
            errors,
        )

    if requested_by_user_id:
        logger.info("inform %s", requested_by_user_id)
        notification_models.Notification.send(
            notification_type=notification_models.Notification.NotificationType.RESENGO_IMPORT_READY,
            notification_result=result,
            message=str(errors),
            user_id=requested_by_user_id,
        )

    return True
